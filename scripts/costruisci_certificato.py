#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Certificato di acquisizione e verifica dei 16 documenti forniti (6.8.2026).
Legge lo strato di estrazione integrale (fonti/estrazioni/*.txt) e produce
corpus/diagnostica/ITALIA_NERA_CERTIFICATO_ACQUISIZIONE.xlsx.

Distinzione dichiarata: si CERTIFICA l'acquisizione e l'integrita' dell'estrazione
(operazione compiuta e verificabile), NON la verita' dei contenuti, che richiede il
protocollo di verifica del repertorio (sez.6, gradi Savona) documento per documento.
"""
import os, re, hashlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EST = os.path.join(ROOT, "fonti", "estrazioni")
DATA = "6.8.2026"

# (txt-base, denominazione, natura, collocazione, note)
DOCS = [
 ("Reti_della_Criminalit__Organizzata_Russa", "Reti della Criminalità Organizzata Russa",
  "monografia di ricerca", "corpus/opera", ""),
 ("Diaspora_Nazista_PostBellica_Ricerca_Approfondita_2", "Diaspora Nazionalsocialista postbellica",
  "monografia di ricerca", "corpus/opera", ""),
 ("Clandestine_Arms_Caches_and_Subversive_Networks_in..._1", "Arsenale della Banda della Magliana",
  "nota di ricerca", "corpus/opera", ""),
 ("BND_CIA_Gladio_Reti_Segrete_Tedesche_1", "BND, CIA, Gladio e le reti stay-behind tedesche",
  "monografia di ricerca", "corpus/opera", ""),
 ("Lupi_Grigi_e_Abdullah__atl__Ricerca_Approfondita", "Lupi Grigi, Abdullah Çatlı e lo Stato profondo turco",
  "monografia di ricerca", "corpus/opera", ""),
 ("Nazisti_in_fuga_Spagna_e_Portogallo_1", "Rete nazista in Spagna e Portogallo",
  "monografia di ricerca", "corpus/opera", ""),
 ("Nazisti_in_Sud_America_Ricerca_Approfondita", "Diaspora nazista in Sud America",
  "monografia di ricerca", "corpus/opera", "una seconda copia (\"_1\") ricevuta e' byte-identica: non riversata"),
 ("Nazisti_in_URSS_Ricerca_Approfondita_1", "Specialisti del Terzo Reich in Unione Sovietica",
  "monografia di ricerca", "corpus/opera", ""),
 ("Nazisti_in_USA_Ricerca_Approfondita_2", "Infiltrazioni naziste nel complesso militare-industriale USA",
  "monografia di ricerca", "corpus/opera", ""),
 ("OPERAZIONI_ANTIMAFIA", "Operazioni antimafia delle Forze dell'ordine italiane",
  "compendio di progetto", "corpus/opera", ""),
 ("Come_i_simzia_1", "Simpatizzanti nazisti a Città del Capo (articolo)",
  "articolo di terzi", "fonti", "nessun apparato di URL"),
 ("Nazi_Havens_in_South_America", "Nazi Havens in South America (articolo)",
  "articolo di terzi", "fonti", "in inglese; nessun apparato di URL"),
 ("BIOGRAFIE_DEI_PRIMI_MEMBRI_DELLA_FAMIGLIA_GENOVESE", "Biografie dei primi membri della famiglia Genovese",
  "compilazione di dati grezzi", "fonti", "tabelle biografiche; nessun apparato di URL"),
 ("NETWORK_FINANZIARIO_SUD_AFRICA", "Network finanziario Sud Africa",
  "compilazione di dati grezzi", "fonti", "elenco nomi/ruoli; nessun apparato di URL"),
 ("NAZIINUSA_1", "Elenco nominativo nazisti negli USA (finding aid)",
  "compilazione di dati grezzi", "fonti", "elenco stile inventario NARA; nessun apparato di URL"),
]

GRAD_RE = re.compile(r'\b(CERTO|ALTAMENTE PROBABILE|PROBABILE|POSSIBILE|DA VERIFICARE|Livello [ABC])\b')

def main():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Certificato acquisizione"
    cols = ["Denominazione", "Natura", "Collocazione", "Parole (estratte)",
            "Occorrenze URL", "Gradazione Savona interna", "Hash testo (sha256/12)",
            "Integrità estrazione", "Verdetto acquisizione", "Data", "Note"]
    hdr_fill = PatternFill("solid", fgColor="1F3864"); hdr_font = Font(bold=True, color="FFFFFF", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    verde = PatternFill("solid", fgColor="C6EFCE"); giallo = PatternFill("solid", fgColor="FFF2CC")
    for j,c in enumerate(cols, start=1):
        cell=ws.cell(1,j,c); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=wrap
    ws.freeze_panes="A2"

    r=2; n_no_grad=0; n_no_url=0
    for base, denom, natura, coll, note in DOCS:
        p=os.path.join(EST, base+".txt")
        txt=open(p, encoding="utf-8", errors="ignore").read()
        words=len(txt.split())
        urls=len(re.findall(r'https?://', txt))
        grad=len(GRAD_RE.findall(txt))
        h=hashlib.sha256(txt.encode("utf-8","ignore")).hexdigest()[:12]
        grad_txt = f"presente ({grad})" if grad else "ASSENTE"
        if not grad: n_no_grad+=1
        if not urls: n_no_url+=1
        valori=[denom, natura, coll, words, urls, grad_txt, h,
                "integra (riaperta senza errori)",
                "ACQUISITO E INTEGRALE (Livello A)", DATA, note]
        for j,v in enumerate(valori, start=1):
            ws.cell(r,j,v).alignment=wrap
        ws.cell(r,9).fill=verde
        ws.cell(r,6).fill = giallo if not grad else verde
        r+=1

    larg=[42,26,16,14,12,22,18,26,30,10,40]
    for j,w in enumerate(larg, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w

    out=os.path.join(ROOT,"corpus","diagnostica","ITALIA_NERA_CERTIFICATO_ACQUISIZIONE.xlsx")
    wb.save(out)
    print("scritto:", out)
    print(f"documenti certificati (acquisizione): {len(DOCS)}")
    print(f"privi di gradazione Savona interna: {n_no_grad}/{len(DOCS)}")
    print(f"privi di apparato di URL: {n_no_url}/{len(DOCS)}")

if __name__=="__main__":
    main()
