#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fase Quattro: collazione del repertorio con i tre censimenti (Guida alle fonti e portale
della Rete italiana degli archivi per non dimenticare; Guía de Archivos de Memoria y
Derechos Humanos en Chile). Produce repertorio/LACUNE_CENSIMENTI.xlsx: le sedi presenti
nei censimenti e ASSENTI dal repertorio (la vera lacuna).

File di servizio (xlsx): esente dalla regola dei numeri per esteso e dal divieto di elenchi.
Data del controllo: 6.8.2026. Gradi Savona: A = istituzione ben attestata su fonti
autorevoli; B = appartenenza/inclusione probabile ma non confermata sul censimento stesso.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DATA = "6.8.2026"

COLONNE = ["Censimento di provenienza", "Denominazione", "Luogo", "Ambito / fondo pertinente",
           "Già nel repertorio?", "Grado", "Note", "Data controllo"]

# Nel repertorio (ARCHIVI_OPERATIVO) figurano, per il fronte cileno: Museo de la Memoria y
# los DDHH e Archivo Nacional de la Memoria cileno; per il fronte italiano della Rete:
# Centro Flamigni, Centro Impastato, Centro Cultura Legalità Democratica (Toscana) e il
# portale stesso. Tutto il resto dei due censimenti è lacuna.
RIGHE = [
 # ---- Guía de Archivos de Memoria y DDHH en Chile ----
 ("Guía Chile (UAH)", "Fundación de Documentación y Archivo de la Vicaría de la Solidaridad",
  "Santiago del Cile", "Assistenza legale alle vittime della dittatura; Monumento Nacional",
  "No", "A", "Nucleo documentario centrale sulla repressione cilena; non censito dal repertorio", DATA),
 ("Guía Chile (UAH)", "Fundación de Ayuda Social de las Iglesias Cristianas (FASIC)",
  "Santiago del Cile", "Assistenza sociale ecclesiale; Monumento Nacional",
  "No", "A", "Assente dal repertorio", DATA),
 ("Guía Chile (UAH)", "Agrupación de Familiares de Detenidos Desaparecidos (AFDD)",
  "Santiago del Cile", "Archivio dei familiari dei desaparecidos; Monumento Nacional",
  "No", "A", "Assente dal repertorio", DATA),
 ("Guía Chile (UAH)", "Corporación de Promoción y Defensa de los Derechos del Pueblo (CODEPU)",
  "Santiago del Cile", "Difesa dei diritti del popolo; fondo presso il Museo de la Memoria; Monumento Nacional",
  "No", "A", "Assente dal repertorio", DATA),
 ("Guía Chile (UAH)", "Comisión Chilena de Derechos Humanos",
  "Santiago del Cile", "Documentazione dei diritti umani; Monumento Nacional",
  "No", "A", "Assente dal repertorio", DATA),
 ("Guía Chile (UAH)", "Fundación PIDEE (Protección a la Infancia Dañada por los Estados de Emergencia)",
  "Santiago del Cile", "Infanzia colpita dalla repressione; fondo presso il Museo de la Memoria; Monumento Nacional",
  "No", "A", "Assente dal repertorio", DATA),
 ("Guía Chile (UAH)", "Archivo Teleanálisis",
  "Santiago del Cile", "Archivio giornalistico audiovisivo della dittatura; Monumento Nacional",
  "No", "A", "Assente dal repertorio", DATA),
 ("Guía Chile (UAH)", "Corporación Parque por la Paz Villa Grimaldi",
  "Santiago del Cile", "Ex centro di detenzione e tortura; archivio orale",
  "No", "A", "Assente dal repertorio", DATA),
 ("Guía Chile (UAH)", "Archivo de Derechos Humanos, Universidad de Chile",
  "Santiago del Cile", "Registro Memoria del Mondo UNESCO (2003)",
  "No", "A", "Assente dal repertorio", DATA),
 ("Guía Chile (UAH)", "Archivo Patrimonial Universidad Alberto Hurtado",
  "Santiago del Cile", "Fondi patrimoniali e di diritti umani (ente curatore della Guía)",
  "No", "A", "Assente dal repertorio", DATA),
 ("Guía Chile (UAH)", "Programa de Derechos Humanos (Subsecretaría de DDHH, Min. Justicia)",
  "Santiago del Cile", "Programma statale sui diritti umani",
  "No", "B", "Inclusione probabile; da confermare sul testo della Guía", DATA),
 # ---- Rete italiana degli archivi per non dimenticare (Guida + portale) ----
 ("Rete italiana (Guida/portale)", "Casa della Memoria di Brescia",
  "Brescia", "Strage di Piazza della Loggia; Associazione familiari delle vittime",
  "No", "A", "Aderente accertato; assente dal repertorio", DATA),
 ("Rete italiana (Guida/portale)", "Fondazione ISEC — Istituto per la storia dell'età contemporanea",
  "Sesto San Giovanni (Milano)", "Storia sociale e politica, movimento operaio",
  "No", "A", "Aderente accertato; assente dal repertorio", DATA),
 ("Rete italiana (Guida/portale)", "Casa della memoria e della storia",
  "Roma", "Associazioni della memoria e della Resistenza",
  "No", "B", "Aderenza probabile; da confermare sull'elenco della Rete", DATA),
 ("Rete italiana (Guida/portale)", "Associazione familiari delle vittime della strage di Piazza Fontana",
  "Milano", "Strage di Piazza Fontana",
  "No", "B", "Aderenza probabile; da confermare", DATA),
 ("Rete italiana (Guida/portale)", "Archivio di Stato di Roma",
  "Roma", "Atti giudiziari e di pubblica sicurezza",
  "No", "B", "Coinvolto nel progetto della Rete; aderenza formale da confermare", DATA),
 ("Rete italiana (Guida/portale)", "[oltre sessanta aderenti complessivi]",
  "Territorio nazionale", "Terrorismo, stragi, eversione, criminalità organizzata",
  "No", "B", "La Rete conta oltre sessanta aderenti: elenco nominale non integralmente "
  "recuperabile in questo ambiente (portale 403). Collazione da completare sul portale/Guida", DATA),
]

def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lacune dai censimenti"
    hdr_fill = PatternFill("solid", fgColor="7F1D1D")   # rosso scuro: sono le lacune
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    gialla = PatternFill("solid", fgColor="FFF2CC")     # grado B: da confermare

    for j, c in enumerate(COLONNE, start=1):
        cell = ws.cell(1, j, c); cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = wrap
    ws.freeze_panes = "A2"

    for r, riga in enumerate(RIGHE, start=2):
        for j, v in enumerate(riga, start=1):
            cell = ws.cell(r, j, v); cell.alignment = wrap
        if riga[5] == "B":
            ws.cell(r, 6).fill = gialla

    larghezze = [26, 46, 22, 40, 16, 8, 46, 12]
    for j, w in enumerate(larghezze, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

    out = "repertorio/LACUNE_CENSIMENTI.xlsx"
    wb.save(out)
    cile = sum(1 for x in RIGHE if x[0].startswith("Guía"))
    ita = sum(1 for x in RIGHE if x[0].startswith("Rete"))
    a = sum(1 for x in RIGHE if x[5] == "A")
    print(f"scritto: {out}")
    print(f"righe-lacuna: {len(RIGHE)}  (Cile {cile}, Italia {ita})")
    print(f"grado A: {a}  |  grado B: {len(RIGHE)-a}")

if __name__ == "__main__":
    main()
