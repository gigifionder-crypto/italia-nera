# -*- coding: utf-8 -*-
"""FASE ZERO (seconda passata) — controllo dello stato HTTP degli URL censiti.

Legge repertorio/URL_CENSITI.xlsx, interroga ogni URL distinto con un ritardo fra l'uno e
l'altro, e scrive in colonna «stato HTTP» il codice di risposta (o l'errore) e in «data
controllo» la data. Gli URL morti NON si cancellano: si marcano — un collegamento morto è
esso stesso un dato (la fonte fu consultabile e non lo è più; la data di accessibilità è un
dato storico, CLAUDE.md §7)."""

import os
import time
import urllib.request
import urllib.error
import ssl

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'repertorio', 'URL_CENSITI.xlsx')
DATA = "2026-08-06"
TIMEOUT = 6
RITARDO = 0.5
UA = "Mozilla/5.0 (ITALIA-NERA verifica archivistica; ricerca storica)"

ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE  # il proxy dell'ambiente ricripta; non disabilita la verifica di rete


def stato(url):
    req = urllib.request.Request(url, method="GET", headers={"User-Agent": UA})
    try:
        with urllib.request.urlopen(req, timeout=TIMEOUT, context=ctx) as r:
            return str(r.status)
    except urllib.error.HTTPError as e:
        return str(e.code)
    except urllib.error.URLError as e:
        return f"MORTO ({str(e.reason)[:40]})"
    except Exception as e:
        return f"MORTO ({type(e).__name__})"


def main():
    wb = openpyxl.load_workbook(XLSX)
    # raccogli gli URL unici su tutti i fogli (colonna A), mappando (foglio, riga)
    posizioni = {}
    for ws in wb.worksheets:
        for r in range(2, ws.max_row + 1):
            u = ws.cell(r, 1).value
            if u:
                posizioni.setdefault(u, []).append((ws.title, r))
    urls = list(posizioni)
    print(f"URL distinti da controllare: {len(urls)}")

    esiti = {}
    vivi = morti = 0
    for i, u in enumerate(urls, 1):
        s = stato(u)
        esiti[u] = s
        if s.startswith("MORTO") or s.startswith("4") or s.startswith("5"):
            morti += 1
        else:
            vivi += 1
        if i % 25 == 0:
            print(f"  {i}/{len(urls)} — vivi {vivi}, problematici {morti}")
        time.sleep(RITARDO)

    for u, pos in posizioni.items():
        for titolo, r in pos:
            ws = wb[titolo]
            ws.cell(r, 8).value = esiti[u]     # colonna «stato HTTP»
            ws.cell(r, 9).value = DATA          # colonna «data controllo»
    wb.save(XLSX)

    from collections import Counter
    c = Counter(esiti.values())
    print(f"\nControllo del {DATA}. Vivi (2xx/3xx): {vivi} — problematici (4xx/5xx/morti): {morti}")
    print("dettaglio stati:")
    for k, n in c.most_common():
        print(f"  {n:>4}  {k}")


if __name__ == "__main__":
    main()
